import json
import os
from pathlib import Path

from flask import Flask, jsonify, request
from flask_cors import CORS
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
DATA_PATHS = [
    BASE_DIR / "data" / "players.json",
    BASE_DIR.parent / "backend-python" / "data" / "players.json",
    BASE_DIR.parent / "backend" / "data" / "players.json",
]
EXCEL_PATHS = [
    BASE_DIR / "FIFA23_official_data.xlsx",
    BASE_DIR / "data" / "FIFA23_official_data.xlsx",
    BASE_DIR.parent / "FIFA23_official_data.xlsx",
    BASE_DIR.parent / "backend-python" / "FIFA23_official_data.xlsx",
    BASE_DIR.parent / "backend-python" / "data" / "FIFA23_official_data.xlsx",
]

app = Flask(__name__)
CORS(app)


def load_players():
    for data_path in DATA_PATHS:
        if data_path.exists():
            with data_path.open("r", encoding="utf-8") as file:
                return json.load(file)

    for excel_path in EXCEL_PATHS:
        if excel_path.exists():
            return excel_to_players(excel_path)

    checked_paths = [str(path) for path in DATA_PATHS + EXCEL_PATHS]
    raise FileNotFoundError(
        "No player data file was found. Deploy either data/players.json or FIFA23_official_data.xlsx. "
        f"Checked paths: {checked_paths}"
    )


def excel_to_players(excel_path):
    wanted_columns = [
        "ID",
        "Name",
        "Age",
        "Nationality",
        "Overall",
        "Potential",
        "Club",
        "Special",
        "Preferred Foot",
        "International Reputation",
        "Weak Foot",
        "Skill Moves",
        "Work Rate",
        "Body Type",
        "Real Face",
        "Joined",
        "Contract Valid Until",
        "Height",
        "Weight",
    ]

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    wanted_indexes = [
        (index, header)
        for index, header in enumerate(headers)
        if header in wanted_columns
    ]
    numeric_columns = {
        "ID",
        "Age",
        "Overall",
        "Potential",
        "Special",
        "International Reputation",
        "Weak Foot",
        "Skill Moves",
    }

    players = []
    seen_ids = set()

    for row in rows:
        player = {}
        for index, header in wanted_indexes:
            value = row[index] if index < len(row) else ""

            if value is None:
                value = ""

            if header in numeric_columns:
                try:
                    value = int(float(value))
                except (TypeError, ValueError):
                    value = 0
            else:
                value = str(value).strip()

            player[header] = value

        if not player.get("Name"):
            continue

        player_id = player.get("ID")
        if player_id and player_id in seen_ids:
            continue
        if player_id:
            seen_ids.add(player_id)

        if not player.get("Club"):
            player["Club"] = "Free Agent"
        if not player.get("Nationality"):
            player["Nationality"] = "Unknown"

        players.append(player)

    workbook.close()
    return players


def filter_players(players):
    search = request.args.get("search", "").strip().lower()
    club = request.args.get("club", "all")
    rating = request.args.get("rating", type=int)
    limit = request.args.get("limit", type=int)

    result = []
    for player in players:
        name = str(player.get("Name", "")).lower()
        player_club = str(player.get("Club", ""))
        overall = int(player.get("Overall") or 0)

        matches_search = not search or search in name
        matches_club = club == "all" or player_club == club
        matches_rating = rating is None or overall >= rating

        if matches_search and matches_club and matches_rating:
            result.append(player)

    if limit:
        result = result[:limit]

    return result


def build_stats(players):
    if not players:
        return {
            "totalPlayers": 0,
            "averageRating": 0,
            "topPlayer": None,
            "bestPotential": None,
            "clubs": [],
            "nationalities": [],
        }

    top_player = max(players, key=lambda player: int(player.get("Overall") or 0))
    best_potential = max(players, key=lambda player: int(player.get("Potential") or 0))
    average_rating = round(
        sum(int(player.get("Overall") or 0) for player in players) / len(players)
    )

    club_stats = {}
    nationality_stats = {}

    for player in players:
        club = player.get("Club") or "Unknown"
        nationality = player.get("Nationality") or "Unknown"
        overall = int(player.get("Overall") or 0)

        if club not in club_stats:
            club_stats[club] = {"club": club, "players": 0, "totalRating": 0}
        club_stats[club]["players"] += 1
        club_stats[club]["totalRating"] += overall

        nationality_stats[nationality] = nationality_stats.get(nationality, 0) + 1

    clubs = []
    for club in club_stats.values():
        clubs.append(
            {
                "club": club["club"],
                "players": club["players"],
                "averageRating": round(club["totalRating"] / club["players"]),
            }
        )

    clubs = sorted(clubs, key=lambda club: club["averageRating"], reverse=True)[:10]

    nationalities = [
        {"nationality": nationality, "players": count}
        for nationality, count in nationality_stats.items()
    ]
    nationalities = sorted(
        nationalities, key=lambda nationality: nationality["players"], reverse=True
    )[:10]

    return {
        "totalPlayers": len(players),
        "averageRating": average_rating,
        "topPlayer": top_player,
        "bestPotential": best_potential,
        "clubs": clubs,
        "nationalities": nationalities,
    }


@app.route("/")
def home():
    return jsonify(
        {
            "message": "Football Player Statistics Analysis API",
            "routes": ["/health", "/api/players", "/api/stats"],
        }
    )


@app.route("/health")
def health():
    json_exists = any(path.exists() for path in DATA_PATHS)
    excel_exists = any(path.exists() for path in EXCEL_PATHS)
    return jsonify(
        {
            "status": "ok",
            "dataFileFound": json_exists or excel_exists,
            "jsonFileFound": json_exists,
            "excelFileFound": excel_exists,
        }
    )


@app.route("/api/players")
@app.route("/players")
def players():
    return jsonify(filter_players(load_players()))


@app.route("/api/stats")
def stats():
    return jsonify(build_stats(load_players()))


@app.errorhandler(Exception)
def handle_error(error):
    return jsonify({"error": "Server error", "message": str(error)}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)
